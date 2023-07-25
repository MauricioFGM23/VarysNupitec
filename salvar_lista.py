import openpyxl

def salvar_lista_em_excel(lista, nome_coluna, nome_arquivo, nome_aba):
    # Abrir o arquivo existente ou criar um novo
    try:
        wb = openpyxl.load_workbook(nome_arquivo)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # Selecionar a planilha desejada ou criar uma nova
    if nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
    else:
        ws = wb.create_sheet(nome_aba)

    # Encontrar a coluna pelo nome ou criar uma nova
    col = None
    for i in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=i).value == nome_coluna:
            col = i # Coluna encontrada
            break
    if col is None: # Coluna não encontrada
        col = ws.max_column + 1 # Criar uma nova coluna
        ws.cell(row=1, column=col).value = nome_coluna # Escrever o nome da coluna

    # Adicionar os elementos da lista na coluna
    row = 2 # Iniciar na segunda linha
    while ws.cell(row=row, column=col).value is not None: # Enquanto a célula não estiver vazia
        row += 1 # Incrementar a linha
    for item in lista: # Iterar sobre a lista e escrever os valores
        ws.cell(row=row, column=col).value = item
        row += 1 # Incrementar a linha

    # Salvar o arquivo
    wb.save(nome_arquivo)

    print(f'Os elementos foram salvos com sucesso na planilha "{nome_coluna}"!')

# # Definindo o nome do arquivo .xlsx e a planilha que contém os dados
# arquivo = "Novo resumos de proteções.xlsx"
# planilha = "SOFTWARE"
# coluna = 'Nº DA PROTEÇÃO'
# n_protecao = ['BR 51 2023 001737 0']
# salvar_lista_em_excel(n_protecao, coluna, arquivo, planilha)